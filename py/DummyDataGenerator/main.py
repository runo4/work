# -*- coding: utf-8 -*-
import os
from datetime import date
from dateutil.relativedelta import relativedelta

from faker import Faker

import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import *
from tkinter import filedialog as tkFD, messagebox


class window(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)

        w1 = Label(text=u"出力ファイルの保存先を選択してください")
        w1.place(x=30, y=45)
        self.t1 = Entry()
        self.t1.place(x=30, y=70, width=300)

        # 参照ボタン
        self.button1 = Button(text=u"参照>>", command=self.open_filedialog)
        self.button1.place(x=340, y=65)

        w2 = Label(text=u"出力するファイル名を入力してください(拡張子なし)")
        w2.place(x=30, y=110)
        self.t2 = Entry()
        self.t2.place(x=30, y=135, width=200)

        w3 = Label(text=u"生成するダミーデータの件数を入力してください(自然数のみ)")
        w3.place(x=30, y=175)
        self.t3 = Entry()
        self.t3.place(x=30, y=200, width=70)

        # エラーメッセージ
        self.w4 = Label(text="")

        # 生成ボタン
        self.button2 = Button(text=u"生成", command=self.confirm_click)
        self.button2.place(x=170, y=280, width=80, height=40)

    # 生成ボタンをクリックしたときのイベント処理
    def confirm_click(self):
        try:
            input_check = False
            folderPath = self.t1.get()
            fileName = self.t2.get()
            countNum = self.t3.get()

            filePath = folderPath + "\\" + fileName + ".xlsx"

            # ラベルメッセージをクリア
            if self.w4.cget("text") != "":
                self.w4.place_forget()

            # フォルダパスもファイル名もともに入力済みである場合
            if folderPath != "" and folderPath is not None and \
                    fileName != "" and fileName is not None:
                # 入力された回数が正の整数である場合
                if countNum.isdigit():
                    # チェックフラグをTrueに
                    input_check = True
                # 入力された回数が正の整数ではない場合
                else:
                    self.w4 = Label(fg="red", text=u"生成件数は正の整数で指定する必要があります")
                    self.w4.place(x=30, y=230)
            # フォルダパスまたはファイル名が未入力
            else:
                self.w4 = Label(fg="red", text=u"フォルダパスとファイル名は共に入力済みである必要があります")
                self.w4.place(x=30, y=230)

            # チェックフラグがTrueの時(入力に不備が無いとき)
            if input_check:
                info_list = self.generate_dummy_data(countNum)
                self.write_excel_file(filePath, info_list)

        except FileNotFoundError:
            messagebox.showerror("エラー", "指定されたディレクトリは存在しません")
        except OSError:
            messagebox.showerror("エラー", "指定されたディレクトリは不正な形式です")

    # 保存先ディレクトリの指定
    def open_filedialog(self):
        iDir = ""
        folderPath = tkFD.askdirectory(initialdir=iDir)
        # ディレクトリ選択時テキストボックスに既に入力されている文字列を削除
        self.t1.delete(0, tk.END)
        self.t1.insert(tk.END, folderPath)

    # fakerによるダミー情報生成メソッド
    def generate_dummy_data(self, count):
        # Fakerのインスタンス作成
        fkgen = Faker("ja_JP")

        # 現在の日付
        date_now = date.today()

        # 個人情報を人数分格納するリスト
        info_list = []

        # 指定回数（count回）繰り返し
        count = int(count)
        for i in range(count):
            # 一人当たりの個人情報が格納されるリスト
            info = []

            # 18歳以上60歳以下となるような生年月日を生成
            birth_date = fkgen.date_of_birth(minimum_age=18, maximum_age=60)
            # 生年月日から年齢の算出
            age = relativedelta(date_now, birth_date)

            # infoリストにランダム生成されたデータを追加
            info.append(i + 1)
            info.append(fkgen.last_name())
            info.append(fkgen.first_name())
            info.append(fkgen.user_name())
            info.append(str(birth_date))
            info.append(age.years)
            info.append(fkgen.zipcode())
            info.append(fkgen.prefecture())
            info.append(fkgen.city())
            info.append(fkgen.town() + fkgen.chome() + fkgen.ban() + fkgen.gou())
            info.append(fkgen.phone_number())
            info.append(fkgen.email())
            info.append(fkgen.company())

            # info_listリストに追加
            info_list.append(info)

        return info_list

    # エクセルファイルへの転記メソッド
    def write_excel_file(self, path, info_list):
        # Excelファイルを新規に作成
        wb = xl.Workbook()
        # シートをアクティブ化
        ws = wb.active

        # 項目名リストを作成
        column_list = ["No.", "苗字", "名前", "ユーザー名",
                       "生年月日", "年齢", "郵便番号", "都道府県",
                       "市区町村", "町名番地", "電話番号",
                       "メールアドレス", "会社名"]

        # リストからExcelファイルに書き込み
        for y, row in enumerate(info_list):
            for x, cell in enumerate(row):
                ws.cell(row=y + 2,
                        column=x + 1,
                        value=info_list[y][x])

        # 列幅自動調整メソッド呼び出し
        self.adjustment_column_width(ws)

        # 項目名転記処理とセル塗りつぶし
        fill = PatternFill(patternType="solid", fgColor="d7eeff")
        for x in range(len(column_list)):
            ws.cell(row=1,
                    column=x + 1,
                    value=column_list[x])
            ws.cell(row=1, column=x + 1).fill = fill

        # 名前を付けて保存
        wb.save(path)

        # 出力ファイルが保存完了した場合
        if os.path.exists(path):
            messagebox.showinfo("出力が完了しました", "ダミーデータファイルの保存が完了しました")
        else:
            messagebox.showerror("出力に失敗しました", "ダミーデータファイルの保存に失敗しました")

    # 列幅自動調整メソッド
    def adjustment_column_width(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width


if __name__ == "__main__":
    root = Tk()

    root.title("ダミーデータ自動生成アプリ")
    root.geometry("420x360")
    root.resizable(width=False, height=False)

    window(root)
    root.mainloop()
